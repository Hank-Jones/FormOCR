import csv
import io
import json
import tempfile
from collections.abc import Iterable, Iterator
from datetime import datetime
from typing import Any, Literal

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session, joinedload

from app.db.models import Form
from app.db.session import SessionLocal, get_db

router = APIRouter(prefix="/export", tags=["export"])

_META_KEYS = ["form_id", "form_type", "status", "created"]
_EXPORT_STATUSES = {
    "approved",
    "cancelled",
    "needs_type",
    "no_template",
    "pending",
    "processing",
    "rejected",
}


class CustomExportPayload(BaseModel):
    columns: list[str] = Field(default_factory=list)
    rows: list[dict[str, Any]] = Field(default_factory=list)


def _query_forms(
    db: Session,
    form_type_id: int | None,
    review_status: str | None,
    since: datetime | None,
):
    q = db.query(Form).options(joinedload(Form.form_type)).order_by(Form.created_at.desc())
    if form_type_id:
        q = q.filter(Form.form_type_id == form_type_id)
    if review_status:
        status = review_status.strip().lower()
        if status not in _EXPORT_STATUSES:
            raise HTTPException(400, "Unsupported review status")
        q = q.filter(Form.review_status == status)
    if since:
        q = q.filter(Form.created_at >= since)
    return q


def _json_object(raw: str | None) -> dict[str, Any]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _stringify_value(value: Any) -> Any:
    if isinstance(value, dict | list):
        return json.dumps(value, ensure_ascii=False)
    return value


def _display_fields(form: Form) -> dict[str, Any]:
    corrected = _json_object(form.corrected_json)
    validated = _json_object(form.validated_json)
    extracted = _json_object(form.extracted_json)
    keys = set(corrected) | set(validated) | set(extracted)
    out: dict[str, Any] = {}
    for key in sorted(keys):
        corrected_value = corrected.get(key)
        validated_value = validated.get(key)
        extracted_value = extracted.get(key)
        if corrected_value is not None and corrected_value != "":
            out[key] = _stringify_value(corrected_value)
        elif validated_value is not None and validated_value != "":
            out[key] = _stringify_value(validated_value)
        elif isinstance(extracted_value, dict) and "text" in extracted_value:
            out[key] = _stringify_value(extracted_value.get("text", ""))
    return out


def _row_data(form: Form) -> dict[str, Any]:
    return {
        "form_id": form.id,
        "form_type": form.form_type.name if form.form_type else "",
        "status": form.review_status,
        "created": form.created_at.isoformat(),
        **_display_fields(form),
    }


def _iter_rows(forms: Iterable[Form]) -> Iterator[dict[str, Any]]:
    for form in forms:
        yield _row_data(form)


def _query_rows(
    db: Session,
    form_type_id: int | None,
    review_status: str | None,
) -> Iterator[dict[str, Any]]:
    forms = _query_forms(db, form_type_id, review_status, None).yield_per(200)
    yield from _iter_rows(forms)


def _ordered_keys_and_count(rows: Iterable[dict[str, Any]]) -> tuple[list[str], int]:
    field_keys: set[str] = set()
    count = 0
    for row in rows:
        count += 1
        for key in row:
            if key not in _META_KEYS:
                field_keys.add(key)
    return [*_META_KEYS, *sorted(field_keys)], count


def _spreadsheet_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    stripped = value.lstrip()
    if stripped and stripped[0] in ("=", "+", "-", "@"):
        return f"'{value}"
    return value


def _stream_query_rows(
    form_type_id: int | None,
    review_status: str | None,
) -> Iterator[dict[str, Any]]:
    db = SessionLocal()
    try:
        yield from _query_rows(db, form_type_id, review_status)
    finally:
        db.close()


def _stream_csv(rows: Iterable[dict[str, Any]], keys: list[str]) -> Iterator[str]:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=keys)
    writer.writerow({key: _spreadsheet_value(key) for key in keys})
    yield buf.getvalue()
    buf.seek(0)
    buf.truncate(0)

    for row in rows:
        writer.writerow({key: _spreadsheet_value(row.get(key, "")) for key in keys})
        yield buf.getvalue()
        buf.seek(0)
        buf.truncate(0)


def _stream_json(rows: Iterable[dict[str, Any]]) -> Iterator[str]:
    yield "["
    first = True
    for row in rows:
        if first:
            first = False
        else:
            yield ","
        yield json.dumps(row, ensure_ascii=False)
    yield "]"


def _custom_columns(rows: list[dict[str, Any]], requested: list[str]) -> list[str]:
    keys: list[str] = []
    for key in requested:
        if key and key not in keys:
            keys.append(key)
    if keys:
        return keys
    for row in rows:
        for key in row:
            if key not in keys:
                keys.append(key)
    return keys


def _columns(rows: list[dict[str, Any]]) -> list[str]:
    keys, _ = _ordered_keys_and_count(rows)
    return keys


def _json_rows(rows: list[dict[str, Any]], keys: list[str]) -> list[dict[str, Any]]:
    return [{key: row.get(key, "") for key in keys} for row in rows]


def _csv_response(rows: list[dict[str, Any]], keys: list[str]) -> Response:
    if not keys:
        return Response(content="", media_type="text/csv")
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=keys)
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in keys})
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=formocr_export.csv"},
    )


def _xlsx_response(rows: list[dict[str, Any]], keys: list[str]) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    if keys:
        ws.append(keys)
        for row in rows:
            ws.append([row.get(key, "") for key in keys])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=formocr_export.xlsx"},
    )


@router.get("/json")
def export_json(
    form_type_id: int | None = None,
    review_status: str | None = Query(None),
    db: Session = Depends(get_db),
):
    forms = _query_forms(db, form_type_id, review_status, None)
    rows = [_row_data(f) for f in forms]
    keys = _columns(rows)
    return _json_rows(rows, keys)


@router.get("/csv")
def export_csv(
    form_type_id: int | None = None,
    review_status: str | None = None,
    db: Session = Depends(get_db),
):
    forms = _query_forms(db, form_type_id, review_status, None)
    rows = [_row_data(f) for f in forms]
    keys = _columns(rows)
    return _csv_response(rows, keys)


@router.get("/xlsx")
def export_xlsx(
    form_type_id: int | None = None,
    review_status: str | None = None,
    db: Session = Depends(get_db),
):
    forms = _query_forms(db, form_type_id, review_status, None)
    rows = [_row_data(f) for f in forms]
    keys = _columns(rows)
    return _xlsx_response(rows, keys)


@router.post("/custom/{format}")
def export_custom(
    payload: CustomExportPayload,
    format: Literal["csv", "xlsx", "json"],
):
    keys = _custom_columns(payload.rows, payload.columns)
    if format == "json":
        return _json_rows(payload.rows, keys)
    if format == "csv":
        return _csv_response(payload.rows, keys)
    return _xlsx_response(payload.rows, keys)
